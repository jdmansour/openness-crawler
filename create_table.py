# {
#   "einrichtung": "International Psychoanalytic University Berlin",
#   "software": "Moodle",
#   "usage_found": true,
#   "reasoning": "Es wird erwähnt, dass Prof. Dr. Phil C. Langer im Jahr 2022 an einer Fortbildung teilgenommen hat, bei der es um 'Teaching Competence – Lehren mit Moodle: Kollaboratives Arbeiten' ging. Dies deutet darauf hin, dass Moodle oder eine auf Moodle basierende Software in der Einrichtung International Psychoanalytic University Berlin genutzt wird, da der Professor an dieser Universität tätig ist und die Fortbildung seine Fähigkeiten im Bereich der Hochschuldidaktik erweitern sollte."
# }
# {
#   "einrichtung": "International Psychoanalytic University Berlin",
#   "software": "Ilias",
#   "usage_found": false,
#   "reasoning": "(No usage found in any document)"
# }
# {
#   "einrichtung": "International Psychoanalytic University Berlin",
#   "software": "OpenOLAT",
#   "usage_found": false,
#   "reasoning": "(No usage found in any document)"
# }

import sys
import os
import pandas as pd
from collections import defaultdict

from utils import parse_json_objects

from openpyxl.formatting import Rule
from openpyxl.formatting.rule import CellIsRule

from openpyxl.styles import Font, PatternFill, Border
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.styles.differential import DifferentialStyle

def main():
    #filename = "results2.jsonlines"
    filename = sys.argv[1]

    # nested dict
    # data[einrichtung][software] = {
    #     "usage_found": "no",
    #     "reasoning": ""
    # }
    data = defaultdict(lambda: defaultdict(lambda: {
        "usage_found": "",
        "reasoning": ""
    }))

    # Parse JSON-Objekte aus der Datei
    json_objects = list(parse_json_objects(filename))
    print(f"Gefundene JSON-Objekte: {len(json_objects)}")
    
    for entry in json_objects:
        print(f"Processing entry: {entry}")
        einrichtung = entry.get("einrichtung", "")
        software = entry.get("software", "")
        usage_found = entry.get("usage_found", False)
        reasoning = entry.get("reasoning", "")

        if einrichtung and software:
            data[einrichtung][software]["usage_found"] = "yes" if usage_found else "no"
            data[einrichtung][software]["reasoning"] = reasoning


    for item in list(data.items())[:5]:
        einrichtung, software_data = item
        print(f"Einrichtung: {einrichtung}")
        for software, details in software_data.items():
            print(f"  Software: {software}, Nutzung: {details['usage_found']}, Begründung: {details['reasoning']}")
        print()

    # Excel-Export
    excel_filename = os.path.splitext(filename)[0] + "_report.xlsx"
    print(f"Erstelle Excel-Bericht: {excel_filename}")
    create_excel_report(data, excel_filename)

def create_excel_report(data, output_filename: str):
    """Erstellt eine Excel-Tabelle mit Einrichtungen als Zeilen und Software-Programmen als Spalten"""
    
    # Sammle alle einzigartigen Software-Programme
    all_software = set()
    for einrichtung_data in data.values():
        all_software.update(einrichtung_data.keys())
    
    all_software = sorted(list(all_software))
    
    # Erstelle DataFrame
    rows = []
    for einrichtung in sorted(data.keys()):
        row = {"Einrichtung": einrichtung}
        
        # Für jede Software füge sowohl Nutzung als auch Begründung hinzu
        for software in all_software:
            if software in data[einrichtung]:
                row[f"{software}"] = data[einrichtung][software]["usage_found"]
                row[f"{software}_Begründung"] = data[einrichtung][software]["reasoning"]
            else:
                row[f"{software}"] = "no"
                row[f"{software}_Begründung"] = "(No usage found in any document)"
        
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    # Speichere als Excel-Datei
    df.to_excel(output_filename, index=False, engine='openpyxl')
    print(f"Excel-Bericht erstellt: {output_filename}")
    

    # Formatiere die Tabelle:
    # Spalte A: Breite 76
    # Spalten B, E, G: Breite 30
    with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Setze Spaltenbreiten
        worksheet.column_dimensions['A'].width = 76
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['E'].width = 30
        worksheet.column_dimensions['G'].width = 30
        
        # Setze Überschriften fett
        for cell in worksheet[1]:
            cell.font = cell.font.copy(bold=True)

        # Bedingte formatierung für Spalten B, D, F:
        # Wenn der inhalt yes ist, hinterlege die Zelle grün
        # Wenn der Inhalt no ist, hinterlege die Zelle rot

        #for col in ['B', 'D', 'F']:
    
        # Standard Excel palette colors (indexed)
        green_fill = PatternFill(start_color="ceeed0", end_color="ceeed0", fill_type="solid")  # Light green
        red_fill   = PatternFill(start_color="f6c9ce", end_color="f6c9ce", fill_type="solid")  # Light red
        red_text = Font(color="8f1b15")  # Dark red text
        green_text = Font(color="285f17")  # Dark green text
        # Area is the whole of B, D, F:
        n = len(df) + 1
        area = f"B2:B{n} D2:D{n} F2:F{n}"

        # for col in ['B', 'D', 'F']:
            # rule = Rule(
            #     type="containsText",
            #     operator="containsText",
            #     text="yes",
            #     dxf=DifferentialStyle(
            #         fill=PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            #     )
            # )

        rule = CellIsRule(operator="equal", formula=['"yes"'], stopIfTrue=True, fill=green_fill, font=green_text)
        worksheet.conditional_formatting.add(area, rule)

        rule = CellIsRule(operator="equal", formula=['"no"'], stopIfTrue=True, fill=red_fill, font=red_text)
        worksheet.conditional_formatting.add(area, rule)
        # workbook.save(output_filename)


    # Zeige eine Vorschau der ersten paar Zeilen
    print("\nVorschau der ersten 3 Zeilen:")
    print(df.head(3).to_string(index=False))

if __name__ == "__main__":
    main()